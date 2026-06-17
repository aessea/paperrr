import re
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from .models import Instance


HEADERS = {
    "index": 0,
    "sku": 1,
    "quantity": 2,
    "volume": 3,
    "weight": 4,
    "category": 5,
    "weight_max": 8,
    "volume_max": 11,
    "sku_style_max": 14,
    "per_sku_quantity_max": 16,
    "priority_class": 17,
}


def _case_metadata(path: Path) -> tuple[str, str, int]:
    match = re.match(r"(?P<scale>small|medium|large)_(?P<num>\d+)_(?P<orders>\d+)_orders_(?P<cabs>\d+)_cabinets", path.stem)
    if not match:
        raise ValueError(f"Cannot parse case metadata from {path.name}")
    scale = match.group("scale")
    name = f"{scale}_{match.group('num')}"
    return name, scale, int(match.group("cabs"))


def load_instance(path: str | Path, config: dict) -> Instance:
    path = Path(path)
    name, scale, n_cabinets = _case_metadata(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    item_ids = []
    sku_raw = []
    cat_raw = []
    volumes = []
    weights = []
    quantities = []
    high_priority = []
    first_caps = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        item_ids.append(int(row[HEADERS["index"]]))
        sku_raw.append(str(row[HEADERS["sku"]]))
        cat_raw.append(str(row[HEADERS["category"]]))
        quantities.append(float(row[HEADERS["quantity"]]))
        volumes.append(float(row[HEADERS["volume"]]))
        weights.append(float(row[HEADERS["weight"]]))
        high_priority.append(1 if int(row[HEADERS["priority_class"]]) == 1 else 0)
        if first_caps is None:
            first_caps = {
                "volume": float(row[HEADERS["volume_max"]]),
                "weight": float(row[HEADERS["weight_max"]]),
                "sku_style": int(row[HEADERS["sku_style_max"]]),
                "per_sku_qty": float(row[HEADERS["per_sku_quantity_max"]]),
            }

    if first_caps is None:
        raise ValueError(f"No data rows found in {path}")

    sku_labels = sorted(set(sku_raw))
    cat_labels = sorted(set(cat_raw))
    sku_index = {v: i for i, v in enumerate(sku_labels)}
    cat_index = {v: i for i, v in enumerate(cat_labels)}

    sku_ids = np.array([sku_index[v] for v in sku_raw], dtype=np.int32)
    category_ids = np.array([cat_index[v] for v in cat_raw], dtype=np.int32)
    high = np.array(high_priority, dtype=np.int8)

    return Instance(
        name=name,
        scale=scale,
        path=str(path),
        n_cabinets=n_cabinets,
        item_ids=np.array(item_ids, dtype=np.int32),
        sku_ids=sku_ids,
        category_ids=category_ids,
        sku_labels=sku_labels,
        category_labels=cat_labels,
        volumes=np.array(volumes, dtype=np.float64),
        weights=np.array(weights, dtype=np.float64),
        quantities=np.array(quantities, dtype=np.float64),
        high_priority=high,
        priority_scores=high.astype(np.float64),
        volume_caps=np.full(n_cabinets + 1, first_caps["volume"], dtype=np.float64),
        weight_caps=np.full(n_cabinets + 1, first_caps["weight"], dtype=np.float64),
        sku_style_caps=np.full(n_cabinets + 1, first_caps["sku_style"], dtype=np.int32),
        per_sku_qty_caps=np.full(n_cabinets + 1, first_caps["per_sku_qty"], dtype=np.float64),
    )
