from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Instance, SolveResult


DETAIL_SHEET = "明细"
SUMMARY_SHEET = "统计"
SUMMARY_HEADERS = [
    "柜号",
    "SKU种类数",
    "品类种类数",
    "高优物品体积",
    "总装载体积",
    "总装载重量",
]


def detail_rows(
    headers: Sequence[object],
    source_rows: Iterable[Sequence[object]],
    assignment: np.ndarray,
) -> tuple[list[object], list[list[object]]]:
    assignment = np.asarray(assignment, dtype=np.int32)
    out_rows = []
    for idx, row in enumerate(source_rows):
        out_rows.append(list(row) + [int(assignment[idx])])
    return list(headers) + ["柜号"], out_rows


def cabinet_summary_rows(instance: Instance, assignment: np.ndarray) -> list[dict[str, float | int]]:
    assignment = np.asarray(assignment, dtype=np.int32)
    if assignment.shape[0] != instance.n_items:
        raise ValueError("Assignment length does not match number of items")

    rows = []
    for cabinet in range(instance.n_cabinets + 1):
        mask = assignment == cabinet
        item_indices = np.where(mask)[0]
        sku_count = int(len(set(int(instance.sku_ids[i]) for i in item_indices)))
        category_count = int(len(set(int(instance.category_ids[i]) for i in item_indices)))
        high_volume = float(np.sum(instance.volumes[mask & (instance.high_priority == 1)]))
        total_volume = float(np.sum(instance.volumes[mask]))
        total_weight = float(np.sum(instance.weights[mask]))
        rows.append(
            {
                "柜号": int(cabinet),
                "SKU种类数": sku_count,
                "品类种类数": category_count,
                "高优物品体积": high_volume,
                "总装载体积": total_volume,
                "总装载重量": total_weight,
            }
        )
    return rows


def write_solution_workbook(
    instance: Instance,
    result: SolveResult,
    output_path: str | Path,
) -> Path:
    if result.assignment is None:
        raise ValueError(f"{instance.name} has no assignment to export")

    input_wb = load_workbook(instance.path, read_only=True, data_only=True)
    input_ws = input_wb.active
    row_iter = input_ws.iter_rows(values_only=True)
    headers = list(next(row_iter))
    source_rows = [list(row) for row in row_iter]
    detail_header, details = detail_rows(headers, source_rows, result.assignment)
    summaries = cabinet_summary_rows(instance, result.assignment)

    wb = Workbook()
    detail_ws = wb.active
    detail_ws.title = DETAIL_SHEET
    detail_ws.append(detail_header)
    for row in details:
        detail_ws.append(row)

    summary_ws = wb.create_sheet(SUMMARY_SHEET)
    summary_ws.append(SUMMARY_HEADERS)
    for row in summaries:
        summary_ws.append([row[header] for header in SUMMARY_HEADERS])

    _format_sheet(detail_ws)
    _format_sheet(summary_ws)
    for col in range(1, len(SUMMARY_HEADERS) + 1):
        summary_ws.column_dimensions[get_column_letter(col)].width = 16
    for col in range(1, len(detail_header) + 1):
        detail_ws.column_dimensions[get_column_letter(col)].width = min(24, max(10, len(str(detail_header[col - 1])) + 2))

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    input_wb.close()
    return output_path


def _format_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
